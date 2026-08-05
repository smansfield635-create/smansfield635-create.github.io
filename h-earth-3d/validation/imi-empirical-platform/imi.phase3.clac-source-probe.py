#!/usr/bin/env python3
"""Inventory and extract a bounded CLAC transcript/metadata package by HTTP range."""
from __future__ import annotations

import argparse
import hashlib
import json
import re
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from remotezip import RemoteZip

ARCHIVE_URL = "https://data.csail.mit.edu/placesaudio/CLAC-Dataset.zip"
README_PATH = "CLAC-Dataset/README.txt"
METADATA_PATH = "CLAC-Dataset/metadata.xlsx"
COOKIE_PREFIX = "CLAC-Dataset/cookie_theft/"
EXTRACTION_LIMIT = 240
MINIMUM_TRANSCRIPTS = 120


def sha256_bytes(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


def canonical_sha256(value: Any) -> str:
    text = json.dumps(value, sort_keys=True, separators=(",", ":"), ensure_ascii=False)
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def classify(name: str) -> list[str]:
    lowered = name.lower()
    labels = []
    if any(token in lowered for token in ("metadata", "demographic", "speaker", "worker")):
        labels.append("METADATA_CANDIDATE")
    if "cookie" in lowered:
        labels.append("COOKIE_THEFT_CANDIDATE")
    if any(lowered.endswith(ext) for ext in (".txt", ".csv", ".json", ".tsv")):
        labels.append("TEXT_OR_TABLE")
    if any(lowered.endswith(ext) for ext in (".png", ".jpg", ".jpeg", ".tif", ".tiff")):
        labels.append("IMAGE")
    if lowered.endswith((".wav", ".flac", ".mp3")):
        labels.append("AUDIO")
    return labels


def json_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def normalize_header(value: Any, index: int) -> str:
    text = str(value).strip() if value is not None else ""
    return text or f"column_{index + 1}"


def parse_metadata(raw: bytes) -> dict[str, Any]:
    workbook = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    sheets = []
    for sheet in workbook.worksheets:
        source_rows = [
            row for row in sheet.iter_rows(values_only=True)
            if any(cell is not None and str(cell).strip() for cell in row)
        ]
        if not source_rows:
            sheets.append({"name": sheet.title, "headers": [], "rows": [], "columnProfiles": []})
            continue
        headers = [normalize_header(value, index) for index, value in enumerate(source_rows[0])]
        rows = []
        for source_row in source_rows[1:]:
            record = {
                headers[index]: json_value(source_row[index] if index < len(source_row) else None)
                for index in range(len(headers))
            }
            if any(value is not None and str(value).strip() for value in record.values()):
                rows.append(record)
        profiles = []
        for header in headers:
            values = [
                str(row[header]).strip()
                for row in rows
                if row.get(header) is not None and str(row.get(header)).strip()
            ]
            unique = sorted(set(values))
            profiles.append({
                "header": header,
                "nonemptyCount": len(values),
                "uniqueCount": len(unique),
                "sampleValues": unique[:12]
            })
        sheets.append({
            "name": sheet.title,
            "headers": headers,
            "rows": rows,
            "columnProfiles": profiles
        })
    return {"sheets": sheets}


def speaker_number(name: str) -> int:
    match = re.search(r"/spk(\d+)_transcript\.txt$", name, flags=re.IGNORECASE)
    if not match:
        raise ValueError(f"CLAC_TRANSCRIPT_NAME_UNEXPECTED:{name}")
    return int(match.group(1))


def normalize_transcript(raw: bytes) -> str:
    text = raw.decode("utf-8", errors="replace").replace("\r\n", "\n").replace("\r", "\n")
    return "\n".join(line.rstrip() for line in text.split("\n")).strip()


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output-dir", required=True)
    parser.add_argument("--clock", default="2026-08-05T19:45:00.000Z")
    args = parser.parse_args()
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    try:
        with RemoteZip(ARCHIVE_URL) as archive:
            infos = archive.infolist()
            inventory = [
                {
                    "name": info.filename,
                    "compressedBytes": info.compress_size,
                    "uncompressedBytes": info.file_size,
                    "crc32": f"{info.CRC:08x}",
                    "classes": classify(info.filename)
                }
                for info in infos
            ]
            names = {entry["name"] for entry in inventory}
            if README_PATH not in names or METADATA_PATH not in names:
                raise RuntimeError("CLAC_REQUIRED_CONTROL_FILES_MISSING")
            readme_raw = archive.read(README_PATH)
            metadata_raw = archive.read(METADATA_PATH)
            transcript_names = sorted(
                (
                    entry["name"] for entry in inventory
                    if entry["name"].startswith(COOKIE_PREFIX)
                    and entry["name"].endswith("_transcript.txt")
                ),
                key=speaker_number
            )
            audio_names = {
                entry["name"] for entry in inventory
                if entry["name"].startswith(COOKIE_PREFIX)
                and entry["name"].endswith(".wav")
            }
            paired_names = [
                name for name in transcript_names
                if name.replace("_transcript.txt", ".wav") in audio_names
            ]
            selected_names = paired_names[:EXTRACTION_LIMIT]
            transcripts = []
            for name in selected_names:
                raw = archive.read(name)
                text = normalize_transcript(raw)
                if len(text) < 20:
                    continue
                number = speaker_number(name)
                transcripts.append({
                    "speakerId": f"spk{number}",
                    "speakerNumber": number,
                    "transcriptPath": name,
                    "audioPath": name.replace("_transcript.txt", ".wav"),
                    "textBytes": len(raw),
                    "textSha256": sha256_bytes(raw),
                    "text": text
                })

        metadata = parse_metadata(metadata_raw)
        if len(transcripts) < MINIMUM_TRANSCRIPTS:
            raise RuntimeError(f"CLAC_MINIMUM_TRANSCRIPTS_NOT_MET:{len(transcripts)}")
        inventory_digest = canonical_sha256(inventory)
        package_body = {
            "schemaVersion": "IMI_PHASE_3_CLAC_PUBLIC_TEXT_METADATA_PACKAGE_v1",
            "observedAt": args.clock,
            "archiveUrl": ARCHIVE_URL,
            "archiveEntries": len(inventory),
            "archiveInventorySha256": inventory_digest,
            "readme": {
                "path": README_PATH,
                "bytes": len(readme_raw),
                "sha256": sha256_bytes(readme_raw),
                "text": readme_raw.decode("utf-8", errors="replace")
            },
            "metadata": {
                "path": METADATA_PATH,
                "bytes": len(metadata_raw),
                "sha256": sha256_bytes(metadata_raw),
                **metadata
            },
            "cookieTheft": {
                "transcriptCountInArchive": len(transcript_names),
                "audioCountInArchive": len(audio_names),
                "pairedTranscriptAudioCount": len(paired_names),
                "extractionLimit": EXTRACTION_LIMIT,
                "extractedUsableTranscriptCount": len(transcripts),
                "transcripts": transcripts
            },
            "boundaries": {
                "fullArchiveDownloaded": False,
                "audioDownloaded": False,
                "transcriptSubsetExtracted": True,
                "metadataExtracted": True,
                "featureExtractionExecuted": False,
                "routeRetuned": False
            }
        }
        package = {**package_body, "packageSha256": canonical_sha256(package_body)}
        candidates = [entry for entry in inventory if entry["classes"]]
        summary = {
            "schemaVersion": "IMI_PHASE_3_CLAC_SOURCE_PROBE_v2",
            "result": "PASS_CLAC_PUBLIC_TEXT_METADATA_PACKAGE_BOUND",
            "observedAt": args.clock,
            "archiveUrl": ARCHIVE_URL,
            "archiveEntries": len(inventory),
            "classifiedCandidates": len(candidates),
            "transcriptCountInArchive": len(transcript_names),
            "audioCountInArchive": len(audio_names),
            "pairedTranscriptAudioCount": len(paired_names),
            "extractedUsableTranscriptCount": len(transcripts),
            "metadataSheetCount": len(metadata["sheets"]),
            "metadataRowCounts": {
                sheet["name"]: len(sheet["rows"]) for sheet in metadata["sheets"]
            },
            "inventorySha256": inventory_digest,
            "packageSha256": package["packageSha256"],
            "error": None,
            "boundaries": package_body["boundaries"]
        }
    except Exception as exc:
        inventory = []
        package = None
        summary = {
            "schemaVersion": "IMI_PHASE_3_CLAC_SOURCE_PROBE_v2",
            "result": "HELD_CLAC_PUBLIC_TEXT_METADATA_PACKAGE_FAILED",
            "observedAt": args.clock,
            "archiveUrl": ARCHIVE_URL,
            "archiveEntries": 0,
            "error": f"{type(exc).__name__}:{exc}",
            "boundaries": {
                "fullArchiveDownloaded": False,
                "audioDownloaded": False,
                "featureExtractionExecuted": False,
                "routeRetuned": False
            }
        }

    (output_dir / "clac-source-probe.v1.json").write_text(
        json.dumps(summary, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
    )
    if inventory:
        (output_dir / "clac-source-inventory.v1.json").write_text(
            json.dumps(inventory, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
        )
    if package is not None:
        (output_dir / "clac-public-text-metadata-package.v1.json").write_text(
            json.dumps(package, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
        )
    print(json.dumps(summary, indent=2, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
